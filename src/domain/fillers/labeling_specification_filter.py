"""标签仕様書-仕様確認書填充器"""

from pathlib import Path
from typing import Any, Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.infrastructure.template_service import ExcelTemplateFiller


class LabelingSpecificationFiller(ExcelTemplateFiller):
    """标签仕様書-仕様確認書填充器"""

    # 填充器写入/修改过的单元格背景色：RGB(115,159,215) -> HEX 0x739FD7
    _FILLED_BG_COLOR = "FF739FD7"  # ARGB format

    def _apply_filled_background(self, cell) -> None:
        """将单元格背景色设置为填充高亮色"""
        cell.fill = PatternFill(fill_type="solid", fgColor=self._FILLED_BG_COLOR)

    def fill_template(
        self,
        template_path: Path,
        parameters: Dict[str, Any],
        output_path: Path,
        language: Optional[str] = None,
    ) -> bool:
        """
        填充标签仕様書-仕様確認書模板

        Args:
            template_path: 模板文件路径
            parameters: 填充参数
            output_path: 输出文件路径

        Returns:
            bool: 是否成功
        """
        try:
            # 优先使用外部传入语言，否则从模板路径提取
            resolved_language = (language or self._extract_language_from_path(template_path)).strip().lower() or "zh"
            # 设置语言（用于空值兜底）
            self._set_language(resolved_language)

            workbook = load_workbook(template_path)
            worksheet = workbook.active

            # 填充字段
            self._fill_fields(worksheet, parameters, resolved_language)

            # 替换其他占位符
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = self._replace_placeholders(cell.value, parameters)

            workbook.save(output_path)
            return True
        except Exception as e:
            print(f"标签仕様書模板填充失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _extract_language_from_path(self, template_path: Path) -> str:
        """从模板路径中提取语言代码"""
        parts = template_path.parts
        for part in parts:
            if part in ['zh', 'ja', 'en']:
                return part
        return 'zh'

    def _fill_fields(self, worksheet, parameters: Dict[str, Any], language: str = 'zh'):
        """填充字段到指定单元格，空值时使用语言对应的兜底文本"""
        missing_text = self._missing_text()

        # theme_no 填入 D5 单元格
        if 'theme_no' in parameters and parameters['theme_no']:
            worksheet['D5'].value = str(parameters['theme_no'])
        else:
            worksheet['D5'].value = missing_text
        self._apply_filled_background(worksheet['D5'])

        # theme_name 填入 M5 单元格
        if 'theme_name' in parameters and parameters['theme_name']:
            worksheet['M5'].value = str(parameters['theme_name'])
        else:
            worksheet['M5'].value = missing_text
        self._apply_filled_background(worksheet['M5'])

        # product_model_name 填入 D7 单元格
        if 'product_model_name' in parameters and parameters['product_model_name']:
            worksheet['D7'].value = str(parameters['product_model_name'])
        else:
            worksheet['D7'].value = missing_text
        self._apply_filled_background(worksheet['D7'])

        # representative_model 填入 G11 单元格 (代表型号)
        if 'representative_model' in parameters and parameters['representative_model']:
            worksheet['G11'].value = str(parameters['representative_model'])
        else:
            worksheet['G11'].value = missing_text
        self._apply_filled_background(worksheet['G11'])

        # product_model 填入 G12 单元格 (商品型式名)
        if 'product_model' in parameters and parameters['product_model']:
            worksheet['G12'].value = str(parameters['product_model'])
        else:
            worksheet['G12'].value = missing_text
        self._apply_filled_background(worksheet['G12'])

        # product_name 填入 G13 单元格
        if 'product_name' in parameters and parameters['product_name']:
            worksheet['G13'].value = str(parameters['product_name'])
        else:
            worksheet['G13'].value = missing_text
        self._apply_filled_background(worksheet['G13'])

        # sales_name 填入 G14 单元格
        if 'sales_name' in parameters and parameters['sales_name']:
            worksheet['G14'].value = str(parameters['sales_name'])
        else:
            worksheet['G14'].value = missing_text
        self._apply_filled_background(worksheet['G14'])

        worksheet['G15'].value = "Intellisense"
        self._apply_filled_background(worksheet['G15'])
        worksheet['G16'].value = "OMRON"
        self._apply_filled_background(worksheet['G16'])

        # target_area 包含 OHC 时填入 G19 单元格, 否则保持空白
        target_area = str(parameters.get('target_area', ''))
        if 'OHC' in target_area.upper():
            worksheet['G19'].value = "OHC提供"
            self._apply_filled_background(worksheet['G19'])
            worksheet['G26'].value = "欧姆龙健康医疗（中国）有限公司"
            self._apply_filled_background(worksheet['G26'])

        # sales_channel 填入 E21 单元格
        if 'sales_channel' in parameters and parameters['sales_channel']:
                worksheet['G21'].value = "400-889-0089"
        else:
                worksheet['G21'].value = "400-770-9988"

        worksheet['G25'].value = "注册证编号/产品技术要求编号"
        self._apply_filled_background(worksheet['G25'])

        # sales_channel：若未提供（None/空）则用兜底文本，避免空白
        if not ('sales_channel' in parameters and str(parameters['sales_channel']).strip()):
            worksheet['G21'].value = missing_text
        self._apply_filled_background(worksheet['G21'])

        if 'address' in parameters and parameters['address']:
            worksheet['G17'].value = parameters['address']
        if 'country' in parameters and parameters['country']:
            worksheet['G18'].value = parameters['country'] + "制造"

        # production_area
        if 'address' in parameters and parameters['address']:
            worksheet['G17'].value = parameters['address']
        else:
            worksheet['G17'].value = missing_text
        self._apply_filled_background(worksheet['G17'])

        if 'country' in parameters and parameters['country']:
            worksheet['G18'].value = parameters['country']
        else:
            worksheet['G18'].value = missing_text
        self._apply_filled_background(worksheet['G18'])
