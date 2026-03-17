"""使用说明书仕样书填充器"""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.infrastructure.template_service import ExcelTemplateFiller

logger = logging.getLogger(__name__)


class UserManualSpecificationFiller(ExcelTemplateFiller):
    """使用说明书仕样书填充器"""

    # 填充器写入/修改过的单元格背景色：RGB(115,159,215) -> HEX 0x739FD7
    _FILLED_BG_COLOR = "FF739FD7"  # ARGB format

    def _apply_filled_background(self, cell) -> None:
        """将单元格背景色设置为填充高亮色"""
        cell.fill = PatternFill(fill_type="solid", fgColor=self._FILLED_BG_COLOR)

    def fill_template(
        self,
        template_path: Path,
        parameters: Dict[str, Any],
        output_path: Path,
        language: Optional[str] = None,
    ) -> bool:
        """
        填充使用说明书仕样书模板

        Args:
            template_path: 模板文件路径
            parameters: 填充参数
            output_path: 输出文件路径

        Returns:
            bool: 是否成功
        """
        non_empty_fields = [k for k, v in parameters.items() if v]
        logger.info("[UserManualSpecificationFiller] 填充字段: %s", non_empty_fields)
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active

            # 设置语言（用于空值兜底）
            self._set_language(language)

            # 填充字段
            self._fill_fields(worksheet, parameters)

            # 替换其他占位符
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = self._replace_placeholders(cell.value, parameters)

            workbook.save(output_path)
            return True
        except Exception as e:
            logger.error("使用说明书仕样书模板填充失败: %s", str(e), exc_info=True)
            return False

    def _fill_fields(self, worksheet, parameters: Dict[str, Any]):
        """填充字段到指定单元格"""
        # 获取空值兜底文本
        missing_text = self._missing_text()

        # theme_no 填入 B19 单元格
        if 'theme_no' in parameters and parameters['theme_no']:
            worksheet['B19'].value = str(parameters['theme_no'])
            self._apply_filled_background(worksheet['B19'])
        else:
            worksheet['B19'].value = missing_text
            self._apply_filled_background(worksheet['B19'])

        # theme_name 填入 D19 单元格
        if 'theme_name' in parameters and parameters['theme_name']:
            worksheet['D19'].value = str(parameters['theme_name'])
            self._apply_filled_background(worksheet['D19'])
        else:
            worksheet['D19'].value = missing_text
            self._apply_filled_background(worksheet['D19'])

        # product_model_name 填入 J19 单元格
        if 'product_model_name' in parameters and parameters['product_model_name']:
            worksheet['J19'].value = str(parameters['product_model_name'])
            self._apply_filled_background(worksheet['J19'])
        else:
            worksheet['J19'].value = missing_text
            self._apply_filled_background(worksheet['J19'])

        # sales_name 填入 B21 单元格
        if 'sales_name' in parameters and parameters['sales_name']:
            worksheet['B21'].value = str(parameters['sales_name'])
            self._apply_filled_background(worksheet['B21'])
        else:
            worksheet['B21'].value = missing_text
            self._apply_filled_background(worksheet['B21'])

        related_file_info: List[Dict[str, Any]] = parameters.get('related_file_info', [])
        mapping: Dict[str, Dict[str, Any]] = {}
        for item in related_file_info:
            for name in str(item.get('short_name', '')).split('|'):
                name = name.strip()
                if name:
                    mapping[name] = item
        row = 25
        while True:
            cell_value = worksheet.cell(row=row, column=1).value  # A列
            if cell_value is None or str(cell_value).strip() == '':
                break
            key = str(cell_value).strip()
            cell_d = worksheet.cell(row=row, column=4)
            cell_g = worksheet.cell(row=row, column=7)
            if key in mapping:
                cell_d.value = str(mapping[key]['file_number'])  # D列
                cell_g.value = str(mapping[key]['version'])  # G列
            else:
                cell_d.value = missing_text
                cell_g.value = missing_text
            self._apply_filled_background(cell_d)
            self._apply_filled_background(cell_g)
            row += 1


