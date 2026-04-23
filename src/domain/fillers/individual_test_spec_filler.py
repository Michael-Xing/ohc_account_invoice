"""个别试验要项书填充器

要求：
- 仅填充指定单元格（含合并单元格的左上角单元格）
- 不影响其它单元格内容与样式
- 对 C37 设置左上对齐，并做 2 级缩进
- 支持动态调整行高以完整展示内容
- 支持在填充位置处理 markdown/HTML 表格，将其转换为 Excel 表格插入（C-P 列）
- 各标题行（C41, C51, C60, C70, C79）不会被覆盖，内容填充从隔行开始
- 标题行间的行数随内容动态调整，不够时插入新行，后续标题行自动下移
- 填充逻辑保持一致（text_block + tables 支持）
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from src.infrastructure.template_service import ExcelTemplateFiller


@dataclass
class BlockConfig:
    """区块配置"""
    title_row: int          # 标题行号
    fill_start_row: int    # 填充起始行号
    param_name: str        # 参数名
    is_last: bool = False  # 是否是最后一个区块
    # 计算结果（动态计算）
    content_rows: int = 0
    available_rows: int = 0
    rows_to_insert: int = 0
    cumulative_insert_rows: int = 0  # 到此区块为止累计插入的行数


class IndividualTestSpecFiller(ExcelTemplateFiller):
    """个别试验要项书填充器（Excel）"""

    # 与 DHFIndexFiller 保持一致的高亮背景色（ARGB）
    _FILLED_BG_COLOR = "FF739FD7"  # RGB(115,159,215)

    def _apply_filled_background(self, cell) -> None:
        """将单元格背景色设置为填充高亮色。"""
        cell.fill = PatternFill(fill_type="solid", fgColor=self._FILLED_BG_COLOR)

    def fill_template(
        self,
        template_path: Path,
        parameters: Dict[str, Any],
        output_path: Path,
        language: Optional[str] = None,
    ) -> bool:
        try:
            # 设置语言（用于空值兜底）
            self._set_language(language)

            workbook = load_workbook(template_path)
            worksheet = workbook.active

            self._fill_fields(worksheet, parameters)

            # 替换其它占位符（如果模板里存在 {xxx} / {{xxx}}）
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = self._replace_placeholders(cell.value, parameters)

            workbook.save(output_path)
            return True
        except Exception as e:
            print(f"个别试验要项书模板填充失败: {str(e)}")
            import traceback

            traceback.print_exc()
            return False

    def _fill_fields(self, worksheet, parameters: Dict[str, Any]) -> None:
        # 获取空值兜底文本
        missing_text = self._missing_text()

        # 1) 基础字段：只写值，不动样式
        mapping = {
            "test_name": "D3",  # 合并后单元格 D3
            "test_number": "K3",  # 合并后单元格 K3
            "theme_no": "D5",  # 合并后单元格 D5
            "product_model": "K5",  # 合并后单元格 K5
            "meas_temperature": "L8",  # 合并后单元格 L8
            "meas_humidity": "N8",  # 合并后单元格 N8
        }

        for key, addr in mapping.items():
            val = parameters.get(key, "")
            if val is None:
                val = ""
            # 空值时填充默认值
            if str(val) == "":
                cell = worksheet[addr]
                cell.value = missing_text
                self._apply_filled_background(cell)
            else:
                cell = worksheet[addr]
                cell.value = str(val)
                self._apply_filled_background(cell)

            # 为 product_model (K5) 动态调整行高
            if key == "product_model" and val:
                self._set_wrap_text_and_adjust_row_height(worksheet, cell)

        # 2) 长文本：需要左上对齐 + 行首缩进两个字符（Excel 缩进）
        test_purpose = parameters.get("test_purpose")
        self._set_text_block_with_row_height(
            worksheet=worksheet,
            cell_addr="C37",
            text=str(test_purpose) if test_purpose else missing_text,
        )

        # 3) 动态区块填充（test_conditions, test_method, others, admission_decision_standard, source）
        self._fill_dynamic_blocks(worksheet, parameters, missing_text)

    def _fill_dynamic_blocks(self, worksheet, parameters: Dict[str, Any], missing_text: str) -> None:
        """
        动态区块填充：根据内容行数动态插入行

        区块定义（原始模板行号）：
        区块                        标题行  填充起始行
        test_conditions            C41    C43
        test_method                C50    C52
        others                     C60    C62
        admission_decision_standard C70    C72
        source                     C80    C82
        test_result                C90    - (边界，不填充)

        核心逻辑：
        1. 保存 test_result 后面到模板原始最大行的内容
        2. 按顺序处理每个区块，计算并插入需要的行
        3. 填充内容
        4. 恢复 test_result 后面的原始内容（只恢复模板原始内容，不覆盖填充内容）
        """
        # 保存 test_result 后面（行90之后）到模板原始最大行的内容
        # 注意：只保存原始模板中的行，不包括后续可能插入的行
        template_original_max_row = 90  # test_result 标题行之后的第一个需要恢复的行
        saved_content = self._save_content_after_row(worksheet, 90, worksheet.max_row)

        # 记录填充开始前的 C-P 列合并单元格
        merged_cells_before_fill = {
            str(mr): mr for mr in worksheet.merged_cells.ranges
            if mr.min_col == 3 and mr.max_col == 16
        }

        # 区块配置（标题行和填充起始行）
        blocks = [
            BlockConfig(title_row=41, fill_start_row=43, param_name="test_conditions"),
            BlockConfig(title_row=50, fill_start_row=52, param_name="test_method"),
            BlockConfig(title_row=60, fill_start_row=62, param_name="others"),
            BlockConfig(title_row=70, fill_start_row=72, param_name="admission_decision_standard"),
            BlockConfig(title_row=80, fill_start_row=82, param_name="source"),
            BlockConfig(title_row=90, fill_start_row=None, param_name="test_result", is_last=True),
        ]

        cumulative_insert_rows = 0

        for i, block in enumerate(blocks):
            if block.param_name == "test_result":
                continue

            next_block = blocks[i + 1]
            next_title_row_adjusted = next_block.title_row + cumulative_insert_rows
            current_fill_start = block.fill_start_row + cumulative_insert_rows
            available_rows = next_title_row_adjusted - current_fill_start - 1

            param_value = parameters.get(block.param_name, "")
            text = str(param_value) if param_value else missing_text
            parts = self._parse_mixed_content(text)
            content_rows = self._calculate_content_rows(parts)

            if content_rows > available_rows:
                rows_to_insert = content_rows - available_rows
                insert_row = next_title_row_adjusted
                # 使用新的插入方法，保护后面的内容
                self._insert_rows_protected(worksheet, insert_row, rows_to_insert)
                cumulative_insert_rows += rows_to_insert
                next_title_row_adjusted = next_block.title_row + cumulative_insert_rows
                available_rows = next_title_row_adjusted - current_fill_start - 1

            cell = worksheet.cell(current_fill_start, 3)
            self._apply_filled_background(cell)
            self._fill_mixed_content_with_tables(worksheet, cell, parts)

        # 收集填充过程中新创建的 C-P 列合并单元格
        new_merged_ranges = []
        for mr in worksheet.merged_cells.ranges:
            if mr.min_col == 3 and mr.max_col == 16:
                range_str = str(mr)
                if range_str not in merged_cells_before_fill:
                    new_merged_ranges.append({
                        'min_row': mr.min_row,
                        'max_row': mr.max_row,
                        'min_col': mr.min_col,
                        'max_col': mr.max_col
                    })

        # 填充完成后，恢复 test_result 后面的原始内容
        # 只恢复原始模板中的行（saved_content 中行号 <= template_original_max_row + rows_inserted）
        # 不覆盖填充区域（行90之前）
        self._restore_content_after_row(
            worksheet, 90, cumulative_insert_rows, saved_content,
            extra_merged_ranges=new_merged_ranges,
            fill_end_row=90  # 填充结束行，低于此行的不恢复
        )

    def _calculate_content_rows(self, parts: List[Dict[str, Any]]) -> int:
        """
        计算内容占用的行数（与 _fill_mixed_content_with_tables 逻辑一致）

        规则：
        - 每个对象前间隔一行（第一个对象不需要）
        - 文本：1行（合并单元格），空文本跳过
        - 图片：1行
        - 表格：表头1行 + 数据行数
        """
        if not parts:
            return 0

        total_rows = 0
        for idx, part in enumerate(parts):
            # 1. 对象前间隔一行（第一个对象不需要）
            if idx > 0:
                total_rows += 1

            if part["type"] == "text":
                text = part.get("content", "").strip()
                # 空文本跳过，不占用行数
                if not text:
                    continue
                total_rows += 1

            elif part["type"] == "image":
                total_rows += 1

            elif part["type"] == "table":
                markdown_text = part.get("content", "")
                if not markdown_text:
                    continue
                headers, rows = self._parse_markdown_table(markdown_text)
                if headers:
                    total_rows += 1 + len(rows)  # 表头 + 数据行

        return total_rows

    def _insert_rows_protected(self, worksheet, insert_at_row: int, num_rows: int) -> None:
        """
        保护性插入行（供 _fill_dynamic_blocks 内部使用）

        直接使用标准 insert_rows，后续由 _restore_content_after_row 恢复

        Args:
            worksheet: 工作表对象
            insert_at_row: 插入位置
            num_rows: 要插入的行数
        """
        if num_rows <= 0:
            return
        worksheet.insert_rows(insert_at_row, amount=num_rows)

    def _save_content_after_row(self, worksheet, after_row: int, max_row: int) -> Dict[str, Any]:
        """
        保存指定行之后的所有内容

        Args:
            worksheet: 工作表对象
            after_row: 保存此行之后的内容
            max_row: 最大行号

        Returns:
            Dict: 保存的内容
        """
        saved = {
            'rows': [],
            'merged_ranges': [],
            'column_dimensions': {},
            'row_dimensions': {}
        }

        # 保存行数据
        for row in range(after_row + 1, max_row + 1):
            row_data = {
                'row': row,
                'cells': {},
                'height': worksheet.row_dimensions[row].height
            }
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                row_data['cells'][col] = {
                    'value': cell.value,
                    'data_type': cell.data_type,
                }
            saved['rows'].append(row_data)

        # 保存合并单元格信息
        for merged_range in list(worksheet.merged_cells.ranges):
            if merged_range.min_row > after_row:
                saved['merged_ranges'].append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col
                })

        # 保存列宽
        for col in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            if col_letter in worksheet.column_dimensions:
                saved['column_dimensions'][col] = worksheet.column_dimensions[col_letter].width

        return saved

    def _restore_content_after_row(
        self, worksheet, after_row: int, rows_inserted: int, saved: Dict[str, Any],
        extra_merged_ranges: List[Dict[str, int]] = None,
        fill_end_row: int = None
    ) -> None:
        """
        恢复保存的内容到正确位置

        Args:
            worksheet: 工作表对象
            after_row: 原分隔行
            rows_inserted: 插入的行数
            saved: 保存的内容
            extra_merged_ranges: 额外需要恢复的合并单元格列表（用于保留填充过程中新创建的合并）
            fill_end_row: 填充结束行，低于此行号的恢复内容将被跳过（用于避免覆盖填充内容）
        """
        if not saved or not saved.get('rows'):
            return

        # 恢复合并单元格（先清除受影响的，但保留填充区域）
        ranges_to_remove = []
        for merged_range in list(worksheet.merged_cells.ranges):
            if merged_range.min_row > after_row:
                # 如果设置了 fill_end_row，跳过填充区域内的合并
                if fill_end_row is not None and merged_range.min_row <= fill_end_row:
                    continue
                ranges_to_remove.append(str(merged_range))
        for range_str in ranges_to_remove:
            try:
                worksheet.unmerge_cells(range_str)
            except Exception:
                pass

        # 恢复行数据（跳过填充区域）
        for row_data in saved['rows']:
            original_row = row_data['row']
            new_row = original_row + rows_inserted

            # 跳过填充区域的行
            if fill_end_row is not None and new_row <= fill_end_row:
                continue

            # 恢复单元格值（跳过合并单元格中的非左上角单元格）
            for col, cell_data in row_data['cells'].items():
                cell = worksheet.cell(new_row, col)
                # 跳过合并单元格中的非左上角单元格
                if isinstance(cell, MergedCell):
                    continue
                cell.value = cell_data['value']

            # 恢复行高
            if row_data.get('height'):
                worksheet.row_dimensions[new_row].height = row_data['height']

        # 恢复合并单元格（跳过填充区域）
        for merged_info in saved['merged_ranges']:
            # 跳过填充区域的合并
            if fill_end_row is not None and merged_info['min_row'] <= fill_end_row:
                continue
            try:
                new_min_row = merged_info['min_row'] + rows_inserted
                new_max_row = merged_info['max_row'] + rows_inserted

                col_letter_start = get_column_letter(merged_info['min_col'])
                col_letter_end = get_column_letter(merged_info['max_col'])
                merged_range_str = f"{col_letter_start}{new_min_row}:{col_letter_end}{new_max_row}"
                worksheet.merge_cells(merged_range_str)
            except Exception:
                pass

        # 恢复额外传入的合并单元格（填充过程中新创建的）
        if extra_merged_ranges:
            for merged_info in extra_merged_ranges:
                try:
                    col_letter_start = get_column_letter(merged_info['min_col'])
                    col_letter_end = get_column_letter(merged_info['max_col'])
                    merged_range_str = f"{col_letter_start}{merged_info['min_row']}:{col_letter_end}{merged_info['max_row']}"
                    worksheet.merge_cells(merged_range_str)
                except Exception:
                    pass

    def _set_wrap_text_and_adjust_row_height(self, worksheet, cell) -> None:
        """设置单元格自动换行并动态调整行高"""
        if cell.alignment:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
                text_rotation=cell.alignment.text_rotation,
                wrap_text=True,
                shrink_to_fit=cell.alignment.shrink_to_fit,
                indent=cell.alignment.indent
            )
        else:
            cell.alignment = Alignment(wrap_text=True)
        self._adjust_row_height_for_text(worksheet, cell.row, cell.column)

    def _adjust_row_height_for_text(self, worksheet, row: int, column: int) -> None:
        """
        根据文本长度自适应调整行高
        
        Args:
            worksheet: 工作表对象
            row: 行号
            column: 列号
        """
        cell = worksheet.cell(row, column)
        text = cell.value
        if not text:
            return
        
        # 获取列宽（字符数），如果未设置则使用默认值
        column_letter = get_column_letter(column)
        column_width = worksheet.column_dimensions[column_letter].width
        if not column_width or column_width == 0:
            column_width = 15  # 默认列宽
        
        # 获取字体大小，如果未设置则使用默认值
        font_size = 11  # 默认字体大小
        if cell.font and cell.font.size:
            font_size = cell.font.size
        
        # 估算每行可容纳的字符数（考虑中文字符占2个位置，英文占1个位置）
        chars_per_line = int(column_width * 1.7)
        if chars_per_line <= 0:
            chars_per_line = 15
        
        # 计算需要的行数
        lines = text.split('\n')
        total_lines = 0
        for line in lines:
            if line:
                line_count = (len(line) + chars_per_line - 1) // chars_per_line
                total_lines += max(1, line_count)
            else:
                total_lines += 1
        
        if total_lines == 0:
            total_lines = 1
        
        # 计算行高（点）
        row_height = total_lines * font_size * 1.5
        
        # 设置最小和最大行高限制
        min_height = font_size * 1.2
        max_height = font_size * 25
        row_height = max(min_height, min(row_height, max_height))
        
        # 设置行高（如果当前行高小于计算出的行高，则更新）
        current_height = worksheet.row_dimensions[row].height
        if not current_height or current_height < row_height:
            worksheet.row_dimensions[row].height = row_height

    def _set_text_block_with_row_height(self, worksheet, cell_addr: str, text: str) -> None:
        """设置长文本块并动态调整行高"""
        if text == "":
            return

        cell = worksheet[cell_addr]
        cell.value = text
        self._apply_filled_background(cell)

        # 设置对齐（左对齐）和换行
        cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            indent=1,
            wrap_text=True,
        )

        # 动态调整行高
        self._adjust_row_height_for_text(worksheet, cell.row, cell.column)

    def _set_test_conditions_with_tables(self, worksheet, cell_addr: str, text: str) -> None:
        """
        设置试验条件文本，支持 markdown 表格转换为 Excel 表格
        内容结构：文本 + 表格 + 文本 + 表格 ...
        
        Args:
            worksheet: 工作表对象
            cell_addr: 起始单元格地址
            text: 要填充的文本
        """
        if text == "":
            return

        # 解析混合内容（文本和表格）
        parts = self._parse_mixed_content(text)
        
        if not parts:
            return
        
        cell = worksheet[cell_addr]
        self._apply_filled_background(cell)

        # 统一由 _fill_mixed_content_with_tables 处理（支持纯文本、表格、图片等）
        self._fill_mixed_content_with_tables(worksheet, cell, parts)

    def _fill_mixed_content_with_tables(
        self, worksheet, first_cell, parts: List[Dict[str, Any]]
    ) -> None:
        """
        填充混合内容（文本+表格+图片）

        结构：
        - 从指定单元格作为起点
        - 每个对象（文本/表格/图片）间隔1个标准行
        - 文本：合并 C-P 列，左对齐，自适应行高
        - 图片：下载后按顺序插入单元格
        - 表格：按 markdown/HTML 语法解析后插入

        Args:
            worksheet: 工作表对象
            first_cell: 起始单元格
            parts: 解析后的内容列表
        """
        current_row = first_cell.row

        for idx, part in enumerate(parts):
            # 1. 对象前间隔一行（第一个对象不需要间隔）
            if idx > 0:
                current_row += 1

            if part["type"] == "text":
                text = part.get("content", "").strip()
                # 剔除前后的空行
                text = self._trim_empty_lines(text)
                if not text:
                    continue
                cell_range = f"C{current_row}:P{current_row}"
                self._merge_and_set_text(worksheet, cell_range, text)
                current_row += 1  # 文本占一行

            elif part["type"] == "image":
                image_url = part.get("content", "")
                if not image_url:
                    continue
                image_path = self._download_image(image_url)
                if image_path:
                    self._insert_image_in_cell(worksheet, current_row, 3, image_path)
                    current_row += 1  # 图片占一行

            elif part["type"] == "table":
                markdown_text = part.get("content", "")
                if not markdown_text:
                    continue
                table_start_row = current_row
                table_rows_count = self._insert_excel_table(worksheet, table_start_row, markdown_text)
                current_row = table_start_row + table_rows_count

    def _trim_empty_lines(self, text: str) -> str:
        """剔除文本前后的空行"""
        if not text:
            print(f"[DEBUG] _trim_empty_lines: 输入为空")
            return ""
        lines = text.split('\n')
        print(f"[DEBUG] _trim_empty_lines: 原始 {len(lines)} 行")
        # 剔除前后空行
        start_idx = 0
        end_idx = len(lines) - 1
        while start_idx <= end_idx and not lines[start_idx].strip():
            start_idx += 1
        while end_idx >= start_idx and not lines[end_idx].strip():
            end_idx -= 1
        if start_idx > end_idx:
            print(f"[DEBUG] _trim_empty_lines: 全部都是空行")
            return ""
        result = '\n'.join(lines[start_idx:end_idx + 1])
        print(f"[DEBUG] _trim_empty_lines: 处理后 {end_idx - start_idx + 1} 行, 结果长度={len(result)}")
        return result

    def _adjust_row_height_for_merged_text(self, worksheet, row: int, text: str) -> None:
        """
        根据文本内容动态调整合并单元格的行高

        Args:
            worksheet: 工作表对象
            row: 行号
            text: 单元格文本
        """
        if not text:
            worksheet.row_dimensions[row].height = 20
            return

        # 获取列宽（C列到P列 = 14列）
        # 使用列字母而不是通过 cell.column_letter 避免 MergedCell 问题
        total_width = 0
        from openpyxl.utils import get_column_letter
        for col_idx in range(3, 17):  # C=3 到 P=16
            col_letter = get_column_letter(col_idx)
            col_width = worksheet.column_dimensions[col_letter].width
            if col_width:
                total_width += col_width
            else:
                total_width += 8  # 默认列宽

        if total_width == 0:
            total_width = 120  # 默认总宽度

        # 获取字体大小
        font_size = 11  # 默认字体大小
        cell = worksheet.cell(row, 3)
        if cell.font and cell.font.size:
            font_size = cell.font.size

        # 估算每行可容纳的字符数（考虑中文字符占2个位置）
        chars_per_line = int(total_width * 1.7)
        if chars_per_line <= 0:
            chars_per_line = 120

        # 计算需要的行数
        lines = text.split('\n')
        total_lines = 0
        for line in lines:
            if line:
                line_count = (len(line) + chars_per_line - 1) // chars_per_line
                total_lines += max(1, line_count)
            else:
                total_lines += 1

        if total_lines == 0:
            total_lines = 1

        # 计算行高
        row_height = total_lines * font_size * 1.5

        # 设置最小和最大行高限制
        min_height = font_size * 1.5
        max_height = font_size * 30
        row_height = max(min_height, min(row_height, max_height))

        worksheet.row_dimensions[row].height = row_height

    def _merge_and_set_text(self, worksheet, cell_range: str, text: str) -> None:
        """
        合并单元格区域并设置文本，根据内容动态调整行高

        Args:
            worksheet: 工作表对象
            cell_range: 单元格范围，如 "C5:P5"
            text: 要设置的文本
        """
        # 取消可能存在的合并
        try:
            worksheet.unmerge_cells(cell_range)
        except (ValueError, KeyError):
            pass

        # 合并单元格
        worksheet.merge_cells(cell_range)

        # 获取合并后的单元格（左上角）
        start_cell_addr = cell_range.split(":")[0]
        cell = worksheet[start_cell_addr]
        cell.value = text
        self._apply_filled_background(cell)

        # 设置对齐
        cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            indent=1,
            wrap_text=True,
        )

        # 获取合并区域的起始行
        start_row = int(cell_range.split(":")[0][1:])
        # 根据内容动态调整行高
        self._adjust_row_height_for_merged_text(worksheet, start_row, text)

    def _insert_excel_table(self, worksheet, start_row: int, markdown_text: str) -> int:
        """
        在指定位置插入 Excel 表格

        Args:
            worksheet: 工作表对象
            start_row: 起始行号
            markdown_text: markdown 或 HTML 格式的表格文本

        Returns:
            int: 插入的表格总行数
        """
        headers, rows = self._parse_markdown_table(markdown_text)
        if not headers:
            return 0

        # 表格列数
        num_cols = len(headers)

        # 确保列数不超过可用列数（C到P = 14列）
        max_cols = 14
        if num_cols > max_cols:
            num_cols = max_cols
            headers = headers[:max_cols]

        # 计算结束行
        total_rows = 1 + len(rows)  # 表头 + 数据行
        end_row = start_row + total_rows - 1

        # 先取消目标区域的合并单元格
        self._unmerge_cells_in_range(worksheet, start_row, end_row, 3, 16)

        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 获取字体大小
        font_size = 11  # 默认字体大小
        cell = worksheet.cell(start_row, 3)
        if cell.font and cell.font.size:
            font_size = cell.font.size

        # 计算列宽（用于行高估算）
        col_widths = []
        for col_idx in range(3, 3 + num_cols):
            col_letter = get_column_letter(col_idx)
            col_width = worksheet.column_dimensions[col_letter].width
            if col_width:
                col_widths.append(col_width)
            else:
                col_widths.append(10)  # 默认列宽

        # 写入表头
        for col_idx, header in enumerate(headers, start=3):
            cell = worksheet.cell(start_row, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 写入数据行
        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx in range(num_cols):
                cell = worksheet.cell(start_row + row_idx, col_idx + 3)
                val = row_data[col_idx] if col_idx < len(row_data) else ""
                cell.value = val
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 设置行高（根据内容动态调整）
        # 表头行高
        header_height = self._calculate_row_height_for_cells(
            headers, col_widths, font_size, is_header=True
        )
        worksheet.row_dimensions[start_row].height = header_height

        # 数据行高
        for row_idx, row_data in enumerate(rows, start=1):
            row_cells = [row_data[i] if i < len(row_data) else "" for i in range(num_cols)]
            data_height = self._calculate_row_height_for_cells(
                row_cells, col_widths, font_size, is_header=False
            )
            worksheet.row_dimensions[start_row + row_idx].height = data_height

        # 应用高亮背景
        for row_idx in range(total_rows):
            for col_idx in range(num_cols):
                cell = worksheet.cell(start_row + row_idx, col_idx + 3)
                self._apply_filled_background(cell)

        return total_rows

    def _calculate_row_height_for_cells(
        self, cells: List[str], col_widths: List[float], font_size: float, is_header: bool = False
    ) -> float:
        """
        根据单元格内容计算需要的行高

        Args:
            cells: 单元格内容列表
            col_widths: 对应列宽列表
            font_size: 字体大小
            is_header: 是否为表头行

        Returns:
            float: 需要的行高
        """
        if not cells:
            return font_size * 2

        # 计算每列每行可容纳的字符数（考虑中文字符占2个位置）
        chars_per_col = [int(w * 1.7) for w in col_widths]

        # 计算每列需要的行数
        max_lines = 1
        for i, cell_text in enumerate(cells):
            if i >= len(chars_per_col):
                break
            if not cell_text:
                continue
            chars_per_line = chars_per_col[i]
            if chars_per_line <= 0:
                chars_per_line = 10
            lines_needed = (len(cell_text) + chars_per_line - 1) // chars_per_line
            max_lines = max(max_lines, lines_needed)

        # 计算行高
        line_height = font_size * 1.5
        row_height = max_lines * line_height

        # 设置最小和最大行高限制
        min_height = font_size * 1.5
        max_height = font_size * 10
        row_height = max(min_height, min(row_height, max_height))

        # 表头行高增加一些
        if is_header:
            row_height = max(row_height, font_size * 2.5)

        return row_height

    def _unmerge_cells_in_range(
        self, worksheet, start_row: int, end_row: int, start_col: int, end_col: int
    ) -> None:
        """
        取消指定行范围内涉及特定列区域的合并单元格
        
        Args:
            worksheet: 工作表对象
            start_row: 起始行
            end_row: 结束行
            start_col: 起始列
            end_col: 结束列
        """
        # 复制一份 ranges，避免迭代时修改集合
        merged_ranges_to_remove = []
        for merged_range in list(worksheet.merged_cells.ranges):
            # 检查合并区域是否与目标区域重叠
            if (merged_range.min_row <= end_row and merged_range.max_row >= start_row and
                merged_range.min_col <= end_col and merged_range.max_col >= start_col):
                merged_ranges_to_remove.append(str(merged_range))
        
        # 取消合并
        for range_str in merged_ranges_to_remove:
            try:
                worksheet.unmerge_cells(range_str)
            except (ValueError, KeyError):
                pass

    def _parse_mixed_content(self, text: str) -> List[Dict[str, Any]]:
        """
        解析混合的 markdown/HTML 内容，识别文本段落、表格和图片

        返回一个列表，每个元素是一个字典：
        - type: 'text', 'table' 或 'image'
        - content: 对应的内容（文本字符串、表格的 markdown/HTML 字符串或图片 URL）
        """
        if not text:
            return []

        parts = []
        current_text_lines = []

        # 先检查是否是 HTML 表格（整个文本就是一个表格）
        if re.search(r'<table[^>]*>.*?</table>', text, re.DOTALL | re.IGNORECASE):
            # 提取所有 HTML 表格
            html_tables = re.findall(r'<table[^>]*>.*?</table>', text, re.DOTALL | re.IGNORECASE)
            if html_tables and len(html_tables) == 1 and text.strip() == html_tables[0].strip():
                # 整个文本就是一个 HTML 表格
                return [{"type": "table", "content": text.strip()}]

        lines = text.splitlines()
        i = 0

        while i < len(lines):
            line = lines[i]
            line_stripped = line.strip()

            # 检查是否是图片（markdown 图片语法或 HTML img 标签）
            image_url = self._extract_image_url(line_stripped)
            if image_url:
                # 先保存之前的文本内容
                if current_text_lines:
                    text_content = "\n".join(current_text_lines)
                    if text_content.strip():
                        parts.append({"type": "text", "content": text_content})
                    current_text_lines = []
                parts.append({"type": "image", "content": image_url})
                i += 1
                continue

            # 检查是否是 HTML 表格的开始（检查是否有 <tr> 或连续的 <td>）
            if re.search(r'<tr[^>]*>.*?</tr>', line_stripped, re.DOTALL | re.IGNORECASE):
                # 收集整个 HTML 表格
                table_start = i
                table_content_parts = []
                table_depth = 0
                while i < len(lines):
                    current_line = lines[i]
                    table_content_parts.append(current_line)
                    # 统计 table 标签
                    table_depth += len(re.findall(r'<table[^>]*>', current_line, re.IGNORECASE))
                    table_depth -= len(re.findall(r'</table>', current_line, re.IGNORECASE))
                    i += 1
                    if table_depth == 0 and '</table>' in ''.join(table_content_parts[-5:]):  # 检查最近几行是否有结束标签
                        break

                table_content = '\n'.join(table_content_parts)

                # 如果之前有文本，保存它
                if current_text_lines:
                    text_content = "\n".join(current_text_lines)
                    if text_content.strip():
                        parts.append({"type": "text", "content": text_content})
                    current_text_lines = []

                # 保存 HTML 表格
                parts.append({"type": "table", "content": table_content})
                continue

            # 检查从当前行开始是否是 markdown 表格的开始
            if i + 1 < len(lines):
                next_line_stripped = lines[i + 1].strip()
                # 检查是否是表格的开始（表头行和分隔符行）
                if ("|" in line_stripped and "|" in next_line_stripped and
                    re.search(r'[-:]+', next_line_stripped)):
                    # 先保存之前的文本内容
                    if current_text_lines:
                        text_content = "\n".join(current_text_lines)
                        if text_content.strip():
                            parts.append({"type": "text", "content": text_content})
                        current_text_lines = []

                    # 收集表格的所有行
                    table_lines = [line]
                    i += 1
                    table_lines.append(lines[i])
                    i += 1

                    # 继续收集表格的数据行（直到遇到非表格行或空行）
                    while i < len(lines):
                        current_line = lines[i]
                        current_line_stripped = current_line.strip()

                        # 如果遇到空行，检查下一行是否是表格的继续
                        if not current_line_stripped:
                            if i + 1 < len(lines):
                                next_line_stripped = lines[i + 1].strip()
                                if "|" in next_line_stripped:
                                    # 下一行是表格行，空行也是表格的一部分
                                    table_lines.append(current_line)
                                    i += 1
                                    continue
                                else:
                                    break
                            else:
                                break

                        # 如果当前行包含 |，可能是表格的数据行
                        if "|" in current_line_stripped:
                            table_lines.append(current_line)
                            i += 1
                        else:
                            break

                    # 保存表格内容
                    table_content = "\n".join(table_lines)
                    if self._is_markdown_table(table_content):
                        parts.append({"type": "table", "content": table_content})
                    else:
                        # 如果不是有效的表格，将其作为文本处理
                        current_text_lines.extend(table_lines)
                    continue

            # 不是表格，作为文本处理
            current_text_lines.append(line)
            i += 1

        # 保存最后的文本内容
        if current_text_lines:
            text_content = "\n".join(current_text_lines)
            if text_content.strip():
                parts.append({"type": "text", "content": text_content})

        return parts

    def _is_markdown_table(self, text: str) -> bool:
        """
        检测文本是否是 markdown 或 HTML 表格格式

        支持格式：
        - Markdown 表格：| 列1 | 列2 | ...
        - HTML 表格：<table>...</table>
        """
        if not text:
            return False

        # 检查是否是 HTML 表格
        if re.search(r'<table', text, re.IGNORECASE):
            return True

        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if len(lines) < 2:
            return False

        header_line = lines[0]
        separator_line = lines[1]

        # 必须包含 | 分隔符
        if "|" not in header_line or "|" not in separator_line:
            return False

        # 表头行应该包含至少一个 |
        header_pipes = header_line.count("|")
        if header_pipes < 2:
            return False

        # 分隔符行应该包含 - 或 : 用于对齐
        if not re.search(r'[-:]+', separator_line):
            return False

        # 验证分隔符行的格式
        separator_clean = separator_line.replace("|", "").replace(" ", "")
        if not re.search(r'[-:]{2,}', separator_clean):
            return False
            return False
        
        return True

    def _parse_markdown_table(self, markdown_text: str) -> Tuple[List[str], List[List[str]]]:
        """
        解析 markdown 或 HTML 表格文本为表头和数据行

        支持格式：
        - Markdown 表格：| 列1 | 列2 | ...
        - HTML 表格：<table>...</table>

        Args:
            markdown_text: 表格文本

        Returns:
            Tuple[List[str], List[List[str]]]: (表头列表, 数据行列表)
        """
        # 检查是否是 HTML 表格
        if re.search(r'<table', markdown_text, re.IGNORECASE):
            return self._parse_html_table(markdown_text)

        lines = [line.strip() for line in markdown_text.splitlines() if line.strip()]
        if len(lines) < 2:
            return [], []

        header_line = lines[0]
        separator_line = lines[1]
        if "|" not in header_line or "|" not in separator_line:
            return [], []

        headers = [cell.strip() for cell in header_line.strip("|").split("|")]
        rows: List[List[str]] = []
        for line in lines[2:]:
            if "|" not in line:
                continue
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if len(cells) < len(headers):
                cells.extend([""] * (len(headers) - len(cells)))
            elif len(cells) > len(headers):
                cells = cells[: len(headers)]
            rows.append(cells)

        return headers, rows

    def _parse_html_table(self, html_text: str) -> Tuple[List[str], List[List[str]]]:
        """
        解析 HTML 表格为表头和数据行

        支持格式：
        - 包含 rowspan/colspan 属性的 HTML 表格
        - 没有 <thead>/<tbody> 标签的扁平结构

        Args:
            html_text: HTML 表格文本

        Returns:
            Tuple[List[str], List[List[str]]]: (表头列表, 数据行列表)
        """
        # 移除 <table> 标签外的内容
        table_match = re.search(r'<table[^>]*>(.*?)</table>', html_text, re.DOTALL | re.IGNORECASE)
        if not table_match:
            return [], []

        table_content = table_match.group(1)

        # 匹配所有 tr 标签
        tr_matches = re.findall(r'<tr[^>]*>(.*?)</tr>', table_content, re.DOTALL | re.IGNORECASE)
        if not tr_matches:
            return [], []

        # 用于存储展开后的所有行
        all_rows: List[List[str]] = []

        # 存储需要向下填充的值（rowspan）
        rowspan_map: List[Dict[int, Tuple[str, int]]] = []  # row_idx -> {col_idx: (value, remaining)}

        for tr_idx, tr_content in enumerate(tr_matches):
            # 匹配所有 td/th 标签
            td_matches = re.findall(r'<t[hd][^>]*(?:/>|>(.*?)</t[hd]>)', tr_content, re.DOTALL | re.IGNORECASE)
            if not td_matches:
                continue

            # 解析当前行的单元格
            current_row: List[str] = []
            col_idx = 0

            # 首先应用之前的 rowspan 值
            if tr_idx < len(rowspan_map):
                for prev_col_idx in range(len(current_row), max(rowspan_map[tr_idx].keys(), default=0) + 1):
                    if prev_col_idx in rowspan_map[tr_idx]:
                        value, remaining = rowspan_map[tr_idx][prev_col_idx]
                        current_row.append(value)
                        # 更新 rowspan 计数
                        if remaining > 1:
                            if tr_idx + 1 >= len(rowspan_map):
                                rowspan_map.append({})
                            rowspan_map[tr_idx + 1][prev_col_idx] = (value, remaining - 1)

            for td_content in td_matches:
                # 获取 rowspan 和 colspan 属性
                rowspan_attr = re.search(r'rowspan=["\']?(\d+)', td_content, re.IGNORECASE)
                colspan_attr = re.search(r'colspan=["\']?(\d+)', td_content, re.IGNORECASE)

                rowspan = int(rowspan_attr.group(1)) if rowspan_attr else 1
                colspan = int(colspan_attr.group(1)) if colspan_attr else 1

                # 获取单元格文本内容
                cell_text = self._strip_html_tags(td_content).strip()

                # 跳过已填充的列（来自 rowspan）
                while col_idx < len(current_row):
                    col_idx += 1

                # 添加单元格内容
                for _ in range(colspan):
                    current_row.append(cell_text)
                    col_idx += 1

                # 处理 rowspan（向下扩展）
                if rowspan > 1:
                    if tr_idx + 1 >= len(rowspan_map):
                        rowspan_map.append({})
                    for offset in range(1, rowspan):
                        if tr_idx + offset >= len(rowspan_map):
                            rowspan_map.append({})
                        rowspan_map[tr_idx + offset][len(current_row) - 1] = (cell_text, rowspan - offset)

            all_rows.append(current_row)

        if not all_rows:
            return [], []

        # 找出最大列数
        max_cols = max(len(row) for row in all_rows)

        # 对齐所有行的列数
        for row in all_rows:
            if len(row) < max_cols:
                row.extend([""] * (max_cols - len(row)))

        # 第一行作为表头，其余作为数据行
        headers = all_rows[0]
        rows = all_rows[1:]

        return headers, rows

    def _strip_html_tags(self, text: str) -> str:
        """移除 HTML 标签，保留文本内容"""
        # 移除所有 HTML 标签
        text = re.sub(r'<[^>]+>', '', text)
        # 解码 HTML 实体
        text = text.replace('&nbsp;', ' ')
        text = text.replace('&lt;', '<')
        text = text.replace('&gt;', '>')
        text = text.replace('&amp;', '&')
        text = text.replace('&quot;', '"')
        return text

    def _extract_image_url(self, line: str) -> Optional[str]:
        """
        从文本行中提取图片 URL

        支持格式：
        - Markdown 图片语法：![alt](url)
        - HTML img 标签：<img src="url"> 或 <img src='url'>
        """
        # Markdown 图片语法：![alt](url)
        md_match = re.search(r'!\[.*?\]\((.*?)\)', line)
        if md_match:
            return md_match.group(1)

        # HTML img 标签：<img src="url"> 或 <img src='url'>
        html_match = re.search(r'<img[^>]+src=["\'](.*?)["\']', line, re.IGNORECASE)
        if html_match:
            return html_match.group(1)

        return None

    def _download_image(self, url: str) -> Optional[Path]:
        """
        下载图片到临时目录

        Args:
            url: 图片 URL

        Returns:
            下载后的图片路径，失败返回 None
        """
        import tempfile
        import urllib.request
        import uuid

        if not url:
            return None

        try:
            # 创建临时目录
            temp_dir = Path(tempfile.gettempdir()) / "ohc_images"
            temp_dir.mkdir(parents=True, exist_ok=True)

            # 生成唯一文件名
            ext = self._get_image_extension(url)
            filename = f"{uuid.uuid4().hex}{ext}"
            file_path = temp_dir / filename

            # 下载图片
            req = urllib.request.Request(
                url,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
            )
            with urllib.request.urlopen(req, timeout=10) as response:
                file_path.write_bytes(response.read())

            return file_path
        except Exception as e:
            print(f"图片下载失败: {url}, 错误: {e}")
            return None

    def _get_image_extension(self, url: str) -> str:
        """从 URL 中提取图片扩展名"""
        # 常见图片扩展名
        extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.svg']
        url_lower = url.lower()
        for ext in extensions:
            if ext in url_lower:
                return ext
        return '.png'  # 默认扩展名

    def _insert_image_in_cell(self, worksheet, row: int, col: int, image_path: Path) -> None:
        """
        在指定单元格插入图片

        Args:
            worksheet: 工作表对象
            row: 行号
            col: 列号（C列 = 3）
            image_path: 图片路径
        """
        from openpyxl.drawing.image import Image as XLImage

        if not image_path or not image_path.exists():
            return

        try:
            # 获取列字母
            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{row}"

            # 取消该行的合并单元格
            self._unmerge_cells_in_range(worksheet, row, row, 3, 16)

            # 加载图片并调整大小以适应列宽
            img = XLImage(str(image_path))

            # 获取列宽和行高（单位：字符/行）
            col_width = worksheet.column_dimensions[col_letter].width or 50
            row_height = worksheet.row_dimensions[row].height or 100

            # 计算图片缩放比例，使其适应单元格
            img_width = img.width
            img_height = img.height

            # 计算宽度缩放比例（基于列宽，约等于字符数 * 7 像素）
            target_width = col_width * 7
            if img_width > target_width:
                scale = target_width / img_width
                img_width = int(img_width * scale)
                img_height = int(img_height * scale)

            # 限制最大高度
            max_height = 400
            if img_height > max_height:
                scale = max_height / img_height
                img_width = int(img_width * scale)
                img_height = max_height

            img.width = img_width
            img.height = img_height

            # 设置行高以容纳图片
            worksheet.row_dimensions[row].height = img_height + 5

            # 添加图片到工作表
            worksheet.add_image(img)
            img.anchor = cell_ref

        except Exception as e:
            print(f"插入图片失败: {image_path}, 错误: {e}")
