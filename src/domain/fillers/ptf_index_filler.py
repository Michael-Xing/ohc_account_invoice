"""PTF INDEX填充器"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

from src.infrastructure.template_service import ExcelTemplateFiller

logger = logging.getLogger(__name__)


class PTFIndexFiller(ExcelTemplateFiller):
    """PTF INDEX填充器"""

    # 填充器写入/修改过的单元格背景色：RGB(115,159,215) -> HEX 0x739FD7
    _FILLED_BG_COLOR = "FF739FD7"  # ARGB format

    def _apply_filled_background(self, cell) -> None:
        """将单元格背景色设置为填充高亮色"""
        cell.fill = PatternFill(fill_type="solid", fgColor=self._FILLED_BG_COLOR)

    def _get_merged_cell_top_left(self, worksheet, row, col):
        """
        获取合并单元格的左上角单元格
        如果单元格不是合并单元格，返回该单元格本身
        """
        cell = worksheet.cell(row, col)
        if isinstance(cell, MergedCell):
            for merged_range in worksheet.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                    return worksheet.cell(merged_range.min_row, merged_range.min_col)
        return cell

    def fill_template(
        self,
        template_path: Path,
        parameters: Dict[str, Any],
        output_path: Path,
        language: Optional[str] = None,
    ) -> bool:
        """
        填充 PTF INDEX 模板。

        执行步骤：
        1. 记录本次填充涉及的非空字段，便于排查。
        2. 设置语言（用于空值兜底文案）。
        3. 加载模板文件并取到活动工作表。
        4. 调用 _fill_data_by_area 根据 target_area 与 file_number_map 执行具体填充。
        5. 将填充后的工作簿保存到输出路径；任何异常都记录日志并返回 False。

        Args:
            template_path: 模板文件路径
            parameters: 填充参数
            output_path: 输出文件路径
            language: 语言代码（用于兜底文案）

        Returns:
            bool: 是否成功
        """
        # 步骤 1：打印非空字段便于定位
        non_empty_fields = [k for k, v in parameters.items() if v]
        logger.info("[PTFIndexFiller] 填充字段: %s", non_empty_fields)
        try:
            # 步骤 2：设置语言（用于空值兜底）
            self._set_language(language)

            # 步骤 3：加载模板并取活动工作表
            workbook = load_workbook(template_path)
            worksheet = workbook.active

            # 步骤 4：执行具体填充逻辑
            self._fill_data_by_area(worksheet, parameters)

            # 步骤 5：保存工作簿
            workbook.save(output_path)
            return True
        except Exception as e:
            logger.error("PTF INDEX模板填充失败: %s", str(e), exc_info=True)
            return False

    # 数据行从 C 列第 19 行开始扫描（C 列为名称列）
    _DATA_START_ROW = 19
    _NAME_COL = 3  # C 列

    @staticmethod
    def _parse_map_entries(file_number_map: List[Any]) -> List[Tuple[str, List[str]]]:
        """
        将 file_number_map 解析为 [(file_number, [name1, name2, ...]), ...]

        执行步骤：
        1. 遍历每一项（兼容 dict 与对象两种形式），取出 file_number 与 short_name。
        2. 将 short_name 按 '|' 分割成多个名字，每个名字去空白后仅保留非空项。
        3. 仅当 file_number 非空且至少有一个有效名字时，才纳入返回列表。
        """
        entries: List[Tuple[str, List[str]]] = []
        for item in file_number_map:
            # 步骤 1：兼容 dict / 对象两种形式取值
            if isinstance(item, dict):
                fn = str(item.get("file_number", "") or "").strip()
                sn = str(item.get("short_name", "") or "")
            else:
                fn = str(getattr(item, "file_number", "") or "").strip()
                sn = str(getattr(item, "short_name", "") or "")
            # 步骤 2：按 '|' 分割 short_name 为多个名字，去空白并过滤空项
            names = [n.strip() for n in sn.split("|") if n.strip()]
            # 步骤 3：过滤掉无效项
            if fn and names:
                entries.append((fn, names))
        return entries

    def _fill_data_by_area(self, worksheet, parameters: Dict[str, Any]) -> None:
        """
        根据 target_area 将 file_number_map 填充到模板的对应单元格。

        执行步骤：
        1. 读取并校验 target_area 参数，按半角逗号分割为多个待匹配的表头文本。
        2. 读取 D15~H15 表头，构建「表头文本 -> 列号」的映射。
        3. 用分割后的 target_area 在表头映射中查找，得到命中的列号集合 matched_cols。
        4. 从第 19 行起向下扫描 C 列，收集 C 列非空的数据行 data_rows；
           C 列为空的行直接跳过，不参与后续匹配与写入。
        5. 解析 file_number_map：将每一项的 short_name 按 '|' 分割成多个名字，
           整理为 [(file_number, [name1, name2, ...]), ...]。
        6. 对每个 matched_cols × data_rows 的单元格：
           a) 若目标格是合并区的从属格（MergedCell），直接跳过。
           b) 取 C 列文本，依次判断是否"包含"任一名字，命中则收集 (首名字, file_number)。
           c) 写入策略（"首名字"取 short_name 以 '|' 分割后的第一个名字）：
              - 无命中：写入兜底文本。
              - 仅命中 1 个 entry：只写 file_number（不带首名字）。
              - 命中 >= 2 个 entry 且首名字全部相同：省略名字前缀，逐行列出 file_number。
              - 命中 >= 2 个 entry 且首名字存在差异：按"首名字"分组，每组先写一行
                "首名字:"，随后逐行列出该组的 file_number；多组之间用空行分隔。
           d) 写入后设置高亮背景色。
        """
        # 步骤 1：读取并校验 target_area，按半角逗号分割
        target_area_raw: str = parameters.get("target_area", "")
        if not target_area_raw:
            return
        target_areas = [a.strip() for a in target_area_raw.split(",") if a.strip()]
        if not target_areas:
            return

        # 步骤 2：读取 D15~H15 表头，构建「表头文本 -> 列号」映射
        header_map: Dict[str, int] = {}
        for col in range(4, 9):  # D=4, E=5, F=6, G=7, H=8
            cell = worksheet.cell(15, col)
            header_value = str(cell.value).strip() if cell.value else ""
            if header_value:
                header_map[header_value] = col

        # 步骤 3：用 target_area 在表头映射中查找，得到命中列号 matched_cols
        matched_cols: List[int] = []
        for area in target_areas:
            col = header_map.get(area)
            if col is not None:
                matched_cols.append(col)
            else:
                logger.warning(
                    "[PTFIndexFiller] target_area '%s' 未在 D15:H15 中匹配到（表头: %s）",
                    area, list(header_map.keys()),
                )
        if not matched_cols:
            return

        # 步骤 4：自第 19 行起向下扫描 C 列，收集 C 列非空的数据行（空行直接跳过）
        data_rows: List[Tuple[int, str]] = []
        max_row = worksheet.max_row or self._DATA_START_ROW
        for row in range(self._DATA_START_ROW, max_row + 1):
            c_cell = worksheet.cell(row, self._NAME_COL)
            c_text = str(c_cell.value).strip() if c_cell.value else ""
            if not c_text:
                continue
            data_rows.append((row, c_text))
        if not data_rows:
            return

        # 步骤 5：解析 file_number_map，将 short_name 按 '|' 分割为多个名字
        file_number_map: List[Any] = parameters.get("file_number_map", []) or []
        entries = self._parse_map_entries(file_number_map)
        missing_text = self._missing_text()

        # 步骤 6：对命中列 × 数据行做"包含"匹配并写入
        for col in matched_cols:
            for row, c_text in data_rows:
                # 6a：遇到合并区从属格（只读）直接跳过
                cell = worksheet.cell(row, col)
                if isinstance(cell, MergedCell):
                    continue

                # 6b：收集所有"包含命中"的 (首名字, file_number)（按 entries 顺序去重）
                matched_items: List[Tuple[str, str]] = []
                for fn, names in entries:
                    if any(name in c_text for name in names):
                        first_name = names[0]
                        pair = (first_name, fn)
                        if pair not in matched_items:
                            matched_items.append(pair)

                # 6c：按命中数量决定写入格式
                if not matched_items:
                    cell.value = missing_text
                elif len(matched_items) == 1:
                    cell.value = matched_items[0][1]
                else:
                    unique_names = {name for name, _ in matched_items}
                    if len(unique_names) == 1:
                        # 全部命中同一首名字 → 省略前缀，仅列出 file_number
                        cell.value = "\n".join(fn for _, fn in matched_items)
                    else:
                        # 名字混合 → 按首名字分组，组内 file_number 顺序写在名字下方
                        groups: Dict[str, List[str]] = {}
                        for name, fn in matched_items:
                            groups.setdefault(name, []).append(fn)
                        blocks: List[str] = []
                        for name, fns in groups.items():
                            blocks.append("\n".join([f"{name}:", *fns]))
                        cell.value = "\n\n".join(blocks)
                # 6d：设置高亮背景色
                self._apply_filled_background(cell)
