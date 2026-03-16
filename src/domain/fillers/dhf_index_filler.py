import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, PatternFill, Protection, Side
from openpyxl.cell.cell import MergedCell

from src.infrastructure.template_service import ExcelTemplateFiller

logger = logging.getLogger(__name__)


class DHFIndexFiller(ExcelTemplateFiller):
    """DHF INDEX填充器"""

    # 该填充器写入/替换过的单元格，统一高亮背景色（ARGB）
    _FILLED_BG_COLOR = "FF739FD7"  # RGB(115,159,215)

    def _apply_filled_background(self, cell) -> None:
        """将单元格背景色设置为填充高亮色。"""
        cell.fill = PatternFill(fill_type="solid", fgColor=self._FILLED_BG_COLOR)
    
    def _adjust_row_height_for_text(self, worksheet, row, text: str, column: int = 3) -> None:
        """
        根据文本长度自适应调整行高
        
        Args:
            worksheet: 工作表对象
            row: 行号
            text: 文本内容
            column: 列号，默认为3（C列）
        """
        if not text:
            return
        
        # 获取单元格
        cell = worksheet.cell(row, column)
        
        # 获取列宽（字符数），如果未设置则使用默认值
        column_letter = cell.column_letter
        column_width = worksheet.column_dimensions[column_letter].width
        if not column_width or column_width == 0:
            column_width = 10  # 默认列宽
        
        # 获取字体大小，如果未设置则使用默认值
        font_size = 11  # 默认字体大小
        if cell.font and cell.font.size:
            font_size = cell.font.size
        
        # 估算每行可容纳的字符数（考虑中文字符占2个位置）
        # 简化处理：假设平均每个字符占1.5个位置（中英文混合）
        chars_per_line = int(column_width * 1.5)
        if chars_per_line <= 0:
            chars_per_line = 15  # 默认值
        
        # 计算需要的行数
        # 考虑文本中可能包含的换行符
        lines = text.split('\n')
        total_lines = 0
        for line in lines:
            if line:
                # 计算这一行需要多少行显示
                line_count = (len(line) + chars_per_line - 1) // chars_per_line
                total_lines += max(1, line_count)
            else:
                total_lines += 1  # 空行也算一行
        
        # 如果只有一行，确保至少显示一行
        if total_lines == 0:
            total_lines = 1
        
        # 计算行高（点）
        # 行高 = 行数 * 字体大小 * 行距系数（1.2-1.5之间，留出一些余量）
        row_height = total_lines * font_size * 1.3
        
        # 设置最小和最大行高限制
        min_height = font_size * 1.2  # 最小行高
        max_height = font_size * 20   # 最大行高（防止过高）
        row_height = max(min_height, min(row_height, max_height))
        
        # 设置行高（如果当前行高小于计算出的行高，则更新）
        current_height = worksheet.row_dimensions[row].height
        if not current_height or current_height < row_height:
            worksheet.row_dimensions[row].height = row_height
    
    def fill_template(
        self,
        template_path: Path,
        parameters: Dict[str, Any],
        output_path: Path,
        language: Optional[str] = None,
    ) -> bool:
        """
        填充DHF INDEX模板
        
        Args:
            template_path: 模板文件路径
            parameters: 填充参数
            output_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        non_empty_fields = [k for k, v in parameters.items() if v]
        logger.info("[DHFIndexFiller] 填充字段: %s", non_empty_fields)
        try:
            # 设置语言（用于空值兜底）
            self._set_language(language)

            # 加载模板文件
            workbook = load_workbook(template_path, data_only=False, keep_vba=False)
            worksheet = workbook.active
            
            # 填充基本信息单元格
            self._fill_basic_info(worksheet, parameters)
            
            # 填充文件列表
            self._fill_file_list(worksheet, parameters)
            
            # 替换其他占位符
            processed_cells = set()
            for row in worksheet.iter_rows():
                for cell in row:
                    # 合并单元格：统一落到左上角单元格处理，并且避免重复处理同一个左上角单元格
                    target_cell = self._get_merged_cell_top_left(worksheet, cell.row, cell.column)
                    cell_key = (target_cell.row, target_cell.column)
                    if cell_key in processed_cells:
                        continue
                    processed_cells.add(cell_key)

                    if target_cell.value and isinstance(target_cell.value, str):
                        original_value = target_cell.value
                        new_value = self._replace_placeholders(original_value, parameters)
                        if new_value != original_value:
                            target_cell.value = new_value
                            self._apply_filled_background(target_cell)
            
            # 保存工作簿
            workbook.save(output_path)
            
            return True
        except Exception as e:
            logger.error("DHF INDEX模板填充失败: %s", str(e), exc_info=True)
            return False

    def _fill_basic_info(self, worksheet, parameters: Dict[str, Any]):
        """填充基本信息到C3-C7单元格"""
        # 获取空值兜底文本
        missing_text = self._missing_text()

        # theme_no: 填充到C3单元格
        if 'theme_no' in parameters and parameters['theme_no']:
            cell_c3 = worksheet['C3']
            cell_c3.value = str(parameters['theme_no'])
            self._apply_filled_background(cell_c3)
        else:
            # 空值时填充默认值
            cell_c3 = worksheet['C3']
            cell_c3.value = missing_text
            self._apply_filled_background(cell_c3)

        # theme_name: 填充到C4单元格
        if 'theme_name' in parameters and parameters['theme_name']:
            cell_c4 = worksheet['C4']
            cell_c4.value = str(parameters['theme_name'])
            self._apply_filled_background(cell_c4)
        else:
            # 空值时填充默认值
            cell_c4 = worksheet['C4']
            cell_c4.value = missing_text
            self._apply_filled_background(cell_c4)

        # product_model: 填充到C5单元格，保留原始值，自适应行高
        if 'product_model' in parameters and parameters['product_model']:
            product_model = str(parameters['product_model'])
            cell_c5 = worksheet['C5']
            cell_c5.value = product_model
            self._apply_filled_background(cell_c5)
            # 设置单元格为自动换行，保留原有对齐方式的其他属性
            if cell_c5.alignment:
                cell_c5.alignment = Alignment(
                    horizontal=cell_c5.alignment.horizontal,
                    vertical=cell_c5.alignment.vertical,
                    text_rotation=cell_c5.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell_c5.alignment.shrink_to_fit,
                    indent=cell_c5.alignment.indent
                )
            else:
                cell_c5.alignment = Alignment(wrap_text=True)
            # 根据文本长度自适应调整行高
            self._adjust_row_height_for_text(worksheet, cell_c5.row, product_model)
        else:
            # 空值时填充默认值
            cell_c5 = worksheet['C5']
            cell_c5.value = missing_text
            self._apply_filled_background(cell_c5)

        # sales_name: 填充到C6单元格，保留原始值，自适应行高
        if 'sales_name' in parameters and parameters['sales_name']:
            sales_name = str(parameters['sales_name'])
            cell_c6 = worksheet['C6']
            cell_c6.value = sales_name
            self._apply_filled_background(cell_c6)
            # 设置单元格为自动换行，保留原有对齐方式的其他属性
            if cell_c6.alignment:
                cell_c6.alignment = Alignment(
                    horizontal=cell_c6.alignment.horizontal,
                    vertical=cell_c6.alignment.vertical,
                    text_rotation=cell_c6.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell_c6.alignment.shrink_to_fit,
                    indent=cell_c6.alignment.indent
                )
            else:
                cell_c6.alignment = Alignment(wrap_text=True)
            # 根据文本长度自适应调整行高
            self._adjust_row_height_for_text(worksheet, cell_c6.row, sales_name)
        else:
            # 空值时填充默认值
            cell_c6 = worksheet['C6']
            cell_c6.value = missing_text
            self._apply_filled_background(cell_c6)

        # stage: 拼接到C7单元格内容的后面
        if 'stage' in parameters and parameters['stage']:
            cell_c7 = worksheet['C7']
            original_value = cell_c7.value or ''
            cell_c7.value = str(original_value) + str(parameters['stage'])
            self._apply_filled_background(cell_c7)

    def _get_merged_cell_top_left(self, worksheet, row, col):
        """
        获取合并单元格的左上角单元格
        如果单元格不是合并单元格，返回该单元格本身
        """
        cell = worksheet.cell(row, col)
        
        # 如果是 MergedCell，找到它所属的合并区域
        if isinstance(cell, MergedCell):
            for merged_range in worksheet.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                    # 返回合并区域的左上角单元格
                    return worksheet.cell(merged_range.min_row, merged_range.min_col)
        
        return cell

    def _copy_cell_style(self, source_cell, target_cell):
        """
        复制源单元格的样式到目标单元格
        
        Args:
            source_cell: 源单元格
            target_cell: 目标单元格
        """
        # 复制字体
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        
        # 复制对齐方式
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
        
        # 复制边框（创建新的 Border 对象，避免 StyleProxy 问题）
        if source_cell.border:
            source_border = source_cell.border
            # 辅助函数：安全地复制 Side 对象
            def copy_side(side_obj):
                if side_obj:
                    return Side(
                        style=getattr(side_obj, 'style', None),
                        color=getattr(side_obj, 'color', None)
                    )
                return None
            
            target_cell.border = Border(
                left=copy_side(getattr(source_border, 'left', None)),
                right=copy_side(getattr(source_border, 'right', None)),
                top=copy_side(getattr(source_border, 'top', None)),
                bottom=copy_side(getattr(source_border, 'bottom', None)),
                diagonal=copy_side(getattr(source_border, 'diagonal', None)),
                diagonal_direction=getattr(source_border, 'diagonal_direction', None),
                outline=getattr(source_border, 'outline', None),
                vertical=copy_side(getattr(source_border, 'vertical', None)),
                horizontal=copy_side(getattr(source_border, 'horizontal', None))
            )
        
        # 复制填充（创建新的 PatternFill 对象，避免 StyleProxy 问题）
        if source_cell.fill and source_cell.fill.patternType:
            source_fill = source_cell.fill
            target_cell.fill = PatternFill(
                fill_type=source_fill.patternType,
                start_color=source_fill.start_color,
                end_color=source_fill.end_color,
                fgColor=source_fill.fgColor,
                bgColor=source_fill.bgColor
            )
        
        # 复制数字格式
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        
        # 复制保护
        if source_cell.protection:
            target_cell.protection = Protection(
                locked=source_cell.protection.locked,
                hidden=source_cell.protection.hidden
            )

    def _fill_file_list(self, worksheet, parameters: Dict[str, Any]):
        """
        填充文件列表
        file_name关键词匹配B29～B196单元格，匹配上后填充number到G列，stage到H列
        填充时会复制B列的样式到G列和H列
        对同名文件进行聚合，同名文件的file_number和stage分别用换行符连接
        """
        if 'file_list' not in parameters or not parameters['file_list']:
            return
        
        file_list = parameters['file_list']
        
        def _normalize_name(name: str) -> str:
            """归一化名称：去除首尾空白/空字符，并将连续空白折叠为单个空格。"""
            if not name:
                return ""
            # 兼容全角空格等
            name = name.replace("\u3000", " ")
            # 兼容零宽空格/BOM 等不可见“空字符”
            name = name.replace("\u200b", "").replace("\ufeff", "")
            name = re.sub(r"\s+", " ", name).strip()
            return name

        def _extract_b_title(cell_text: str) -> str:
            """提取B列用于匹配的文本：去掉首尾空白和空字符后做全量匹配。"""
            if not cell_text:
                return ""
            return _normalize_name(str(cell_text))

        # 按“别名(由short_name按|拆分)”分组聚合
        # key为拆分后的单个名称，value为列表，每个元素是(file_number, stage)元组
        file_groups: Dict[str, List[tuple]] = {}
        
        for file_item in file_list:
            # 处理file_item可能是字典或对象的情况
            if isinstance(file_item, dict):
                file_name = str(file_item.get('short_name', ''))
                number = str(file_item.get('file_number', ''))
                stage = str(file_item.get('stage', ''))
            else:
                # 假设是对象，尝试获取属性
                file_name = str(getattr(file_item, 'short_name', ''))
                number = str(getattr(file_item, 'file_number', ''))
                stage = str(getattr(file_item, 'stage', ''))
            
            if not file_name:
                continue

            # short_name支持用“|”分隔多个名字；任意一个名字命中B列括号前内容即视为匹配
            # 例如：short_name="个装箱图纸|包装箱图纸" 可命中 "个装箱图纸(备注信息)" / "包装箱图纸（备注）"
            aliases = [_normalize_name(x) for x in str(file_name).split("|")]
            aliases = [x for x in aliases if x]
            if not aliases:
                continue

            for alias in aliases:
                if alias not in file_groups:
                    file_groups[alias] = []
                file_groups[alias].append((number, stage))
        
        # 遍历B29～B196单元格，进行关键词匹配
        for row_num in range(29, 197):  # B29到B196
            cell_b = worksheet.cell(row_num, 2)  # B列
            # 获取B列的实际单元格（如果是合并单元格，获取左上角的单元格）
            cell_b_actual = self._get_merged_cell_top_left(worksheet, row_num, 2)
            cell_b_value = str(cell_b_actual.value) if cell_b_actual.value else ''
            b_title = _extract_b_title(cell_b_value)
            if not b_title:
                continue
            
            number_stage_pairs = file_groups.get(b_title)
            if not number_stage_pairs:
                continue

            # 匹配成功，聚合同名文件的file_number和stage
            # 分别用换行符连接，保持对应关系
            # 确保同一条记录的file_number和stage在拼接后的字符串位置顺序一致
            numbers = []
            stages = []
            for number, stage in number_stage_pairs:
                # 即使值为空也添加，保持位置对应关系（用于多行对齐）
                numbers.append(number if number else '')
                stages.append(stage if stage else '')
            
            # 用换行符连接（保留空行占位，确保对齐，不做rstrip）
            aggregated_number = '\n'.join(numbers) if numbers else ''
            aggregated_stage = '\n'.join(stages) if stages else ''
            
            # 获取G列和H列的左上角单元格（如果是合并单元格）
            cell_g = self._get_merged_cell_top_left(worksheet, row_num, 7)  # G列
            cell_h = self._get_merged_cell_top_left(worksheet, row_num, 8)  # H列
            
            # 填充聚合后的值
            if aggregated_number:
                cell_g.value = aggregated_number
            if aggregated_stage:
                cell_h.value = aggregated_stage
            
            # 复制B列的样式到G列和H列
            self._copy_cell_style(cell_b_actual, cell_g)
            self._copy_cell_style(cell_b_actual, cell_h)

            # 标记该填充器写入过的单元格背景色
            if aggregated_number:
                self._apply_filled_background(cell_g)
            if aggregated_stage:
                self._apply_filled_background(cell_h)
            
            # 设置自动换行，以便正确显示多行内容
            if cell_g.alignment:
                cell_g.alignment = Alignment(
                    horizontal=cell_g.alignment.horizontal,
                    vertical=cell_g.alignment.vertical,
                    text_rotation=cell_g.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell_g.alignment.shrink_to_fit,
                    indent=cell_g.alignment.indent
                )
            else:
                cell_g.alignment = Alignment(wrap_text=True)
            
            if cell_h.alignment:
                cell_h.alignment = Alignment(
                    horizontal=cell_h.alignment.horizontal,
                    vertical=cell_h.alignment.vertical,
                    text_rotation=cell_h.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell_h.alignment.shrink_to_fit,
                    indent=cell_h.alignment.indent
                )
            else:
                cell_h.alignment = Alignment(wrap_text=True)
            
            # 根据文本长度自适应调整行高
            # 分别计算G列和H列需要的行高，取较大值
            if aggregated_number:
                self._adjust_row_height_for_text(worksheet, row_num, aggregated_number, column=7)  # G列
            if aggregated_stage:
                self._adjust_row_height_for_text(worksheet, row_num, aggregated_stage, column=8)  # H列


