"""产品环境评估要项书/结果书填充器"""

from pathlib import Path
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

from src.infrastructure.template_service import ExcelTemplateFiller


class ProductEnvironmentAssessmentFiller(ExcelTemplateFiller):
    """产品环境评估要项书/结果书填充器"""
    
    def fill_template(self, template_path: Path, parameters: Dict[str, Any], output_path: Path) -> bool:
        """
        填充产品环境评估要项书/结果书模板
        
        Args:
            template_path: 模板文件路径
            parameters: 填充参数
            output_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
            
            # 1. 填充简单拼接字段
            self._fill_simple_fields(worksheet, parameters)
            
            # 2. 填充表格数据
            self._fill_table_data(worksheet, parameters)
            
            # 替换其他占位符
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = self._replace_placeholders(cell.value, parameters)
            
            workbook.save(output_path)
            return True
        except Exception as e:
            print(f"产品环境评估模板填充失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _fill_simple_fields(self, worksheet, parameters: Dict[str, Any]):
        """填充简单拼接字段"""
        # theme_no 拼接到 B5 单元格内容后面
        if 'theme_no' in parameters and parameters['theme_no']:
            cell_b5 = worksheet['B5']
            original_value = cell_b5.value or ''
            cell_b5.value = str(original_value) + str(parameters['theme_no'])
        
        # product_model_name 拼接到 B7 单元格内容后面
        if 'product_model_name' in parameters and parameters['product_model_name']:
            cell_b7 = worksheet['B7']
            original_value = cell_b7.value or ''
            cell_b7.value = str(original_value) + str(parameters['product_model_name'])
        
        # product_name 拼接到 I5 单元格内容后面
        if 'product_name' in parameters and parameters['product_name']:
            cell_i5 = worksheet['I5']
            original_value = cell_i5.value or ''
            cell_i5.value = str(original_value) + str(parameters['product_name'])
        
        # production_area 拼接到 I6 单元格内容后面
        if 'production_area' in parameters and parameters['production_area']:
            cell_i6 = worksheet['I6']
            original_value = cell_i6.value or ''
            cell_i6.value = str(original_value) + str(parameters['production_area'])
    
    def _fill_table_data(self, worksheet, parameters: Dict[str, Any]):
        """
        填充表格数据
        表头行：21行
        表尾行：26行（初始）
        数据行：22-25行（初始4行）
        """
        # 表头行和表尾行
        HEADER_ROW = 21
        FOOTER_ROW = 27
        START_DATA_ROW = 22
        INITIAL_DATA_ROWS = 5  # 初始数据行数（22-26行）
        
        # 解析数据：按'/'分割
        product_model_list = []
        sales_name_list = []
        target_area_list = []
        
        if 'product_model' in parameters and parameters['product_model']:
            product_model_list = [item.strip() for item in str(parameters['product_model']).split('/') if item.strip()]
        
        if 'sales_name' in parameters and parameters['sales_name']:
            sales_name_list = [item.strip() for item in str(parameters['sales_name']).split('/') if item.strip()]
        
        # target_area 不分割，直接作为单个值使用
        target_area_value = ''
        if 'target_area' in parameters and parameters['target_area']:
            target_area_value = str(parameters['target_area']).strip()
        
        # 计算需要的数据行数（取product_model和sales_name列表的最大长度）
        data_count = max(len(product_model_list), len(sales_name_list))
        
        if data_count == 0:
            return  # 没有数据，直接返回
        
        # target_area 为所有行填充相同的值
        target_area_list = [target_area_value] * data_count
        
        # 获取表头样式
        header_style = self._get_header_style(worksheet, HEADER_ROW)
        
        # 获取现有数据行行高（取22-25行的平均行高或第一个非空行高）
        row_height = self._get_data_row_height(worksheet, START_DATA_ROW, INITIAL_DATA_ROWS)
        
        # 计算需要插入的行数
        rows_to_insert = max(0, data_count - INITIAL_DATA_ROWS)
        
        # 如果需要插入行，在表尾行之前插入
        if rows_to_insert > 0:
            worksheet.insert_rows(FOOTER_ROW, amount=rows_to_insert)
            # 插入行后，表尾行位置会下移
            # 但这里我们不需要更新FOOTER_ROW，因为后续操作不涉及表尾行
        
        # 填充数据
        for i in range(data_count):
            current_row = START_DATA_ROW + i
            
            # 获取数据（如果列表长度不足，使用空字符串）
            product_model = product_model_list[i] if i < len(product_model_list) else ''
            sales_name = sales_name_list[i] if i < len(sales_name_list) else ''
            target_area = target_area_list[i] if i < len(target_area_list) else ''
            
            # 字段1：B～C合并，空字符填充
            # 先取消可能存在的跨行合并，确保只合并单行
            cell_b = worksheet.cell(current_row, 2)  # B列
            # 检查并取消可能存在的跨行合并（只取消跨行的，不取消单行的）
            merged_ranges_to_remove = []
            for merged_range in list(worksheet.merged_cells.ranges):
                # 检查是否涉及B或C列，且是跨行合并（min_row != max_row）
                if (merged_range.min_col <= 3 and merged_range.max_col >= 2 and 
                    merged_range.min_row <= current_row <= merged_range.max_row and
                    merged_range.min_row != merged_range.max_row):
                    merged_ranges_to_remove.append(str(merged_range))
            # 安全地取消合并，捕获可能的KeyError
            for range_str in merged_ranges_to_remove:
                try:
                    worksheet.unmerge_cells(range_str)
                except (KeyError, ValueError):
                    # 如果取消合并失败（单元格不存在等），忽略错误继续执行
                    pass
            cell_b.value = ''
            merge_range_bc = f'B{current_row}:C{current_row}'
            worksheet.merge_cells(merge_range_bc)
            self._apply_merged_cell_style(worksheet, merge_range_bc, header_style)
            
            # 字段2：D～H合并，product_model填充
            cell_d = worksheet.cell(current_row, 4)  # D列
            cell_d.value = product_model
            merge_range_dh = f'D{current_row}:H{current_row}'
            worksheet.merge_cells(merge_range_dh)
            self._apply_merged_cell_style(worksheet, merge_range_dh, header_style)
            
            # 字段3：I～J合并，sales_name填充
            cell_i = worksheet.cell(current_row, 9)  # I列
            cell_i.value = sales_name
            merge_range_ij = f'I{current_row}:J{current_row}'
            worksheet.merge_cells(merge_range_ij)
            self._apply_merged_cell_style(worksheet, merge_range_ij, header_style)
            
            # 字段4：K～M合并，target_area填充
            cell_k = worksheet.cell(current_row, 11)  # K列
            cell_k.value = target_area
            merge_range_km = f'K{current_row}:M{current_row}'
            worksheet.merge_cells(merge_range_km)
            self._apply_merged_cell_style(worksheet, merge_range_km, header_style)
            
            # 设置行高
            if row_height:
                worksheet.row_dimensions[current_row].height = row_height
        
        # 设置表尾行和表后第一行的行高，表后第二行设置为二倍行高
        if row_height:
            # 计算实际表尾行位置（如果插入了行，表尾行会下移）
            actual_footer_row = FOOTER_ROW + rows_to_insert
            # 设置表尾行和表后第一行的行高（与表行高一致）
            for i in range(0, 2):
                row_num = actual_footer_row + i
                # 先访问单元格，确保行存在
                _ = worksheet.cell(row_num, 1)
                # 获取或创建行维度对象并设置行高
                # 注意：在openpyxl中，设置height会自动管理customHeight属性
                worksheet.row_dimensions[row_num].height = row_height
            # 设置表后第二行的行高（二倍行高）
            second_row_after_footer = actual_footer_row + 2
            _ = worksheet.cell(second_row_after_footer, 1)
            worksheet.row_dimensions[second_row_after_footer].height = row_height * 2
    
    def _get_header_style(self, worksheet, header_row: int) -> Dict[str, Any]:
        """获取表头样式"""
        style = {
            'font': None,
            'alignment': None,
            'border': None,
            'fill': None
        }
        
        # 尝试从表头行的各个单元格获取样式
        for col in range(2, 14):  # B到M列
            cell = worksheet.cell(header_row, col)
            if cell.font:
                style['font'] = cell.font
            if cell.alignment:
                style['alignment'] = cell.alignment
            if cell.border:
                style['border'] = cell.border
            if cell.fill and cell.fill.patternType:
                style['fill'] = cell.fill
            # 找到一个有样式的单元格就退出
            if any(style.values()):
                break
        
        return style
    
    def _get_data_row_height(self, worksheet, start_row: int, row_count: int) -> float:
        """获取现有数据行行高"""
        heights = []
        for i in range(row_count):
            row_num = start_row + i
            height = worksheet.row_dimensions[row_num].height
            if height:
                heights.append(height)
        
        # 返回平均行高，如果没有则返回None
        if heights:
            return sum(heights) / len(heights)
        return None
    
    def _apply_merged_cell_style(self, worksheet, merge_range: str, header_style: Dict[str, Any]):
        """
        为合并单元格应用样式
        合并单元格后，需要为合并区域的所有边界单元格设置边框，确保底部边框完整
        """
        # 获取合并区域的边界
        min_col, min_row, max_col, max_row = range_boundaries(merge_range)
        
        # 获取合并后的单元格（左上角）
        merged_cell = worksheet.cell(min_row, min_col)
        
        # 应用字体和对齐样式到合并单元格
        if header_style.get('font'):
            original_font = header_style['font']
            merged_cell.font = Font(
                name=original_font.name,
                size=(original_font.size - 1) if original_font.size else None,
                bold=original_font.bold,
                italic=original_font.italic,
                vertAlign=original_font.vertAlign,
                underline=original_font.underline,
                strike=original_font.strike,
                color=original_font.color
            )
        
        if header_style.get('alignment'):
            original_alignment = header_style['alignment']
            merged_cell.alignment = Alignment(
                horizontal=original_alignment.horizontal,
                vertical=original_alignment.vertical,
                text_rotation=original_alignment.text_rotation,
                wrap_text=original_alignment.wrap_text,
                shrink_to_fit=original_alignment.shrink_to_fit,
                indent=original_alignment.indent
            )
        
        # 为合并区域的所有边界单元格设置边框
        thin_side = Side(style='thin')
        
        # 顶部边框：为所有列设置顶部边框
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(min_row, col)
            current_border = cell.border
            if current_border:
                cell.border = Border(
                    left=current_border.left or thin_side,
                    right=current_border.right or thin_side,
                    top=thin_side,
                    bottom=current_border.bottom or thin_side
                )
            else:
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thin_side,
                    bottom=thin_side
                )
        
        # 底部边框：为所有列设置底部边框（关键修复）
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(max_row, col)
            current_border = cell.border
            if current_border:
                cell.border = Border(
                    left=current_border.left or thin_side,
                    right=current_border.right or thin_side,
                    top=current_border.top or thin_side,
                    bottom=thin_side
                )
            else:
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thin_side,
                    bottom=thin_side
                )
        
        # 左侧边框：为所有行设置左侧边框
        for row in range(min_row, max_row + 1):
            cell = worksheet.cell(row, min_col)
            current_border = cell.border
            if current_border:
                cell.border = Border(
                    left=thin_side,
                    right=current_border.right or thin_side,
                    top=current_border.top or thin_side,
                    bottom=current_border.bottom or thin_side
                )
            else:
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thin_side,
                    bottom=thin_side
                )
        
        # 右侧边框：为所有行设置右侧边框
        for row in range(min_row, max_row + 1):
            cell = worksheet.cell(row, max_col)
            current_border = cell.border
            if current_border:
                cell.border = Border(
                    left=current_border.left or thin_side,
                    right=thin_side,
                    top=current_border.top or thin_side,
                    bottom=current_border.bottom or thin_side
                )
            else:
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thin_side,
                    bottom=thin_side
                )
        
        # 填充：与表头一致（如果有）
        if header_style.get('fill'):
            merged_cell.fill = header_style['fill']
    