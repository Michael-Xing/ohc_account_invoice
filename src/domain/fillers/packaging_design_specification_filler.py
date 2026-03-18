"""包装设计仕样书填充器"""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.infrastructure.template_service import ExcelTemplateFiller

logger = logging.getLogger(__name__)


class PackagingDesignSpecificationFiller(ExcelTemplateFiller):
    """包装设计仕样书填充器"""

    # 填充器写入/修改过的单元格背景色：RGB(115,159,215) -> HEX 0x739FD7
    _FILLED_BG_COLOR = "FF739FD7"  # ARGB format

    def _apply_filled_background(self, cell) -> None:
        """将单元格背景色设置为填充高亮色"""
        cell.fill = PatternFill(fill_type="solid", fgColor=self._FILLED_BG_COLOR)

    def _extract_stage_order(self, stage: Any) -> int:
        """提取阶段中的 DR 数字，未匹配时返回 -1。"""
        match = re.search(r"DR\s*(\d+)", str(stage or ""), re.IGNORECASE)
        return int(match.group(1)) if match else -1

    def _build_related_file_mapping(
        self, related_file_info: List[Dict[str, Any]]
    ) -> Dict[str, Dict[str, Any]]:
        """按 short_name 聚合，保留 DR 数字最大的记录。"""
        mapping: Dict[str, Dict[str, Any]] = {}
        for item in related_file_info:
            key = str(item.get("short_name", "")).strip()
            if not key:
                continue

            current_item = mapping.get(key)
            if current_item is None or self._extract_stage_order(
                item.get("stage")
            ) >= self._extract_stage_order(current_item.get("stage")):
                mapping[key] = item

        return mapping

    def fill_template(
        self,
        template_path: Path,
        parameters: Dict[str, Any],
        output_path: Path,
        language: Optional[str] = None,
    ) -> bool:
        """
        填充包装设计仕样书模板

        Args:
            template_path: 模板文件路径
            parameters: 填充参数
            output_path: 输出文件路径

        Returns:
            bool: 是否成功
        """
        non_empty_fields = [k for k, v in parameters.items() if v]
        logger.info("[PackagingDesignSpecificationFiller] 填充字段: %s", non_empty_fields)
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active

            # 优先使用外部传入语言，否则从模板路径提取
            language = (language or self._extract_language_from_path(template_path)).strip().lower() or "zh"
            # 设置语言（用于空值兜底）
            self._set_language(language)

            # 填充字段
            self._fill_fields(worksheet, parameters, language)

            # 替换其他占位符
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = self._replace_placeholders(cell.value, parameters)

            workbook.save(output_path)
            return True
        except Exception as e:
            logger.error("包装设计仕样书模板填充失败: %s", str(e), exc_info=True)
            return False

    def _extract_language_from_path(self, template_path: Path) -> str:
        """从模板路径中提取语言代码"""
        parts = template_path.parts
        # 路径格式通常是: .../excel/zh/文件名.xlsx 或 .../excel/ja/文件名.xlsx
        for part in parts:
            if part in ['zh', 'ja', 'en']:
                return part
        return 'zh'  # 默认返回中文

    def _fill_fields(self, worksheet, parameters: Dict[str, Any], language: str = 'zh'):
        """填充字段到指定单元格"""
        # 获取空值兜底文本
        missing_text = self._missing_text()

        # theme_no 填入 C21 单元格
        if 'theme_no' in parameters and parameters['theme_no']:
            worksheet['C21'].value = str(parameters['theme_no'])
            self._apply_filled_background(worksheet['C21'])
        else:
            worksheet['C21'].value = missing_text
            self._apply_filled_background(worksheet['C21'])

        # theme_name 填入 E21 单元格
        if 'theme_name' in parameters and parameters['theme_name']:
            worksheet['E21'].value = str(parameters['theme_name'])
            self._apply_filled_background(worksheet['E21'])
        else:
            worksheet['E21'].value = missing_text
            self._apply_filled_background(worksheet['E21'])

        # product_model_name 填入 L21 单元格
        if 'product_model_name' in parameters and parameters['product_model_name']:
            worksheet['L21'].value = str(parameters['product_model_name'])
            self._apply_filled_background(worksheet['L21'])
        else:
            worksheet['L21'].value = missing_text
            self._apply_filled_background(worksheet['L21'])

        # sales_name 填入 C23 单元格
        if 'sales_name' in parameters and parameters['sales_name']:
            worksheet['C23'].value = str(parameters['sales_name'])
            self._apply_filled_background(worksheet['C23'])
        else:
            worksheet['C23'].value = missing_text
            self._apply_filled_background(worksheet['C23'])

        related_file_info: List[Dict[str, Any]] = parameters.get('related_file_info', [])
        mapping = self._build_related_file_mapping(related_file_info)
        row = 27
        while True:
            cell_value = worksheet.cell(row=row, column=2).value  # B列
            if cell_value is None or str(cell_value).strip() == '':
                break
            key = str(cell_value).strip()
            cell_e = worksheet.cell(row=row, column=5)
            if key in mapping:
                cell_e.value = str(mapping[key]['file_number'])  # E列
            else:
                cell_e.value = missing_text
            self._apply_filled_background(cell_e)
            row += 1



